import sys
sys.setrecursionlimit(5000)
block_cipher = None

# main part
from pathlib import Path
from xlsxwriter import Workbook
from os import chdir, path, listdir

#Get folder, regular expression, change folder, etc
import re
import os

#opening the result
from openpyxl import load_workbook

# writing to the new one
from csv import reader, writer

#date
import datetime

#get the directory of py file
resultDir = Path(__file__).resolve().parent

#get the py file if needed
resultPyFile = Path(__file__).resolve()
#current date and time
today = datetime.date.today()
    
#Create Excel File
def createResult():
    chdir(resultDir)
    # Create New XLS File
    if path.exists('Billing Result '+ str(today) + '.xlsx'):
        print("Result already exist!")
    else:
        newXls = Workbook('Billing Result '+ str(today) + '.xlsx')
        # 1st Pln
        transaksi = newXls.add_worksheet('1st PLN')
        #Transaksi Columns 
        transaksi.write('A1', 'ID PEL')
        transaksi.write('B1', 'Kode Rek')
        transaksi.write('C1', 'Nama Pelanggan')
        transaksi.write('D1', 'Subsidi')
        transaksi.write('E1', 'Golongan')
        transaksi.write('F1', 'Daya')
        transaksi.write('G1', 'Faktor Kali Meter')
        transaksi.write('H1', 'Stand Awal LWBP') 
        transaksi.write('I1', 'Stand Akhir LWBP') 
        transaksi.write('J1', 'kWh LWBP') 
        transaksi.write('K1', 'Stand Awal WBP') 
        transaksi.write('L1', 'Stand Akhir WBP') 
        transaksi.write('M1', 'kWh WBP') 
        transaksi.write('N1', 'Stand Awal kVArh')
        transaksi.write('O1', 'Stand Akhir kVArh')
        transaksi.write('P1', 'kVArh')
        transaksi.write('Q1', 'Tarif LWBP')
        transaksi.write('R1', 'Tarif WBP')
        transaksi.write('S1', 'Tarif kVArh')
        transaksi.write('T1', 'Tunggakan Sebelumnya')
        transaksi.write('U1', 'BP(Biaya Penyambungan)')
        transaksi.write('V1', 'Tagihan Susulan')
        transaksi.write('W1', 'P2TL')
        transaksi.write('X1', 'UJL(Uang Jaminan Listrik)')
        transaksi.write('Y1', 'Angsuran Lainnya')
        transaksi.write('Z1', 'Biaya EMIN') 
        transaksi.write('AA1', 'Total Cicilan') 
        transaksi.write('AB1', 'Rupiah TTL Terpakai')
        transaksi.write('AC1', 'Rupiah Kompensasi')
        transaksi.write('AD1', 'PPJ*')
        transaksi.write('AE1', 'PPN*')
        transaksi.write('AF1', 'Rupiah Jasa Layanan')
        transaksi.write('AG1', 'Total Tagihan')
        transaksi.write('AH1', 'Status Pembayaran')
        transaksi.write('AI1', 'Data Insertion Date')
        #column width
        transaksi.set_column(0, 29, 15)
        # Batam Sheet
        batam = newXls.add_worksheet('Batam')
        # Batam Columns
        batam.write('A1', 'ID PEL')
        batam.write('B1', 'Nama Pelanggan')
        batam.write('C1', 'Nomor Rek/Billing Number')
        batam.write('D1', 'Rekening Bulan/ Billing Month')
        batam.write('E1', 'Golongan')
        batam.write('F1', 'Daya')
        batam.write('G1', 'Faktor Kali Meter')
        batam.write('H1', 'Status')
        batam.write('I1', 'Tagihan Sebelumnya')
        batam.write('J1', 'Pembayaran')
        batam.write('K1', 'Tagihan Sekarang/Current Balance')
        batam.write('L1', 'Tariff Adjustment')
        batam.write('M1', 'Biaya Beban')
        batam.write('N1', 'Biaya KWh')
        batam.write('O1', 'Kelebihan kVArh')
        batam.write('P1', 'Materai')
        batam.write('Q1', 'Tambahan Pajak')
        batam.write('R1', 'Late Charge')
        batam.write('S1', 'Transformer Charge')
        batam.write('T1', 'Cap Charge')
        batam.write('U1', 'Persentase PPJ')
        batam.write('V1', 'PPJ in Rupiah')
        batam.write('W1', 'Angsuran')
        batam.write('X1', 'Kompensasi/Quality Reduction')
        batam.write('Y1', 'Lain-lain')
        batam.write('Z1', 'Tagihan Sekarang / Total Amount Due')
        batam.write('AA1', 'Sisa Angsuran')
        batam.write('AB1', 'Tgl KWh Before')
        batam.write('AC1', 'TglKWh After')
        batam.write('AD1', 'Posisi Kwh Before')
        batam.write('AE1', 'Posisi Kwh After')
        batam.write('AF1', 'Blok 1 Pemakaian KWH')
        batam.write('AG1', 'Blok 1 Rupiah Pemakaian KWh')
        batam.write('AH1', 'Blok 1 Perhitungan Pemakaian KWh')
        batam.write('AI1', 'Blok 2 Pemakaian KWH')
        batam.write('AJ1', 'Blok 2 Rupiah Pemakaian KWh')
        batam.write('AK1', 'Blok 2 Perhitungan Pemakaian KWh')
        batam.write('AL1', 'Data Insertion Date')
        
        # column width
        batam.set_column(0, 29, 15)
    
        # mandiri sheet
        mandiri = newXls.add_worksheet('Mandiri')
        mandiri.write('A1', 'ID PEL')
        mandiri.write('B1', 'Nama File')
        mandiri.write('C1', 'Tanggal')
        mandiri.write('D1', 'Jam')
        mandiri.write('E1', 'Golongan')
        mandiri.write('F1', 'Daya')
        mandiri.write('G1', 'Tagihan Rp')
        mandiri.write('H1', 'No. Ref')
        mandiri.write('I1', 'BL/TH')
        mandiri.write('J1', 'Stand Meter')
        mandiri.write('K1', 'Biaya Admin')
        mandiri.write('L1', 'Total Tagihan')
        mandiri.write('M1', 'Data Insertion Date')
        #column width
        mandiri.set_column(0, 29, 15)
        
        #pln 2 sheet
        pln2 = newXls.add_worksheet('2nd PLN')
        pln2.write('A1', 'Nama')
        pln2.write('B1', 'IDPEL')
        pln2.write('C1', 'Rekening Bulan')
        pln2.write('D1', 'Golongan')
        pln2.write('E1', 'Daya')
        pln2.write('F1', 'Faktor Kali Meter')
        pln2.write('G1', 'Jam Nyala')
        pln2.write('H1', 'Stand Akhir Tanggal') 
        pln2.write('I1', 'LWBP Akhir')
        pln2.write('J1', 'WBP Akhir') 
        pln2.write('K1', 'Stand Awal Tanggal')
        pln2.write('L1', 'LWBP Awal')
        pln2.write('M1', 'WBP Awal') 
        pln2.write('N1', 'Stand Meter Reset(Tanggal)')
        pln2.write('O1', 'Stand Meter Reset LWBP')
        pln2.write('P1', 'Stand Meter Cabut(Tanggal)')
        pln2.write('Q1', 'Stand Meter Cabut LWBP')
        pln2.write('R1', 'Pemakaian kWh Total') 
        pln2.write('S1', 'Biaya Beban') 
        pln2.write('T1', 'Pemakaian kWh LWBP') 
        pln2.write('U1', 'Biaya Pemk LWBP') 
        pln2.write('V1', 'Subtotal LWBP') 
        pln2.write('W1', 'Pemakaian kWh WBP')
        pln2.write('X1', 'Biaya Pemk WBP')
        pln2.write('Y1', 'Subtotal WBP')
        pln2.write('Z1', 'Kelebihan kVarh')
        pln2.write('AA1', 'Biaya kVarh')
        pln2.write('AB1', 'Subtotal kVarh')
        pln2.write('AC1', 'Total Pemakapan')
        pln2.write('AD1', 'PTL Bruto')
        pln2.write('AE1', 'Rupiah Kompensasi')
        pln2.write('AF1', 'PPJ Total')
        pln2.write('AG1', 'RPTTL')
        pln2.write('AH1', 'Tagihan Lainnya')
        pln2.write('AI1', 'Restitusi')
        pln2.write('AJ1', 'RPTL') 
        pln2.write('AK1', 'PPN') 
        pln2.write('AL1', 'Total Tagihan')
        pln2.write('AM1', 'PPJ(Persentase)')
        pln2.write('AN1', 'PPJ')
        pln2.write('AO1', 'Jumlah Tagihan')
        pln2.write('AP1', 'Batas')
        pln2.write('AQ1', 'Status')
        pln2.write('AR1', 'Tanggal')
        pln2.write('AS1', 'Biaya Keterlambatan')
    
        #corrupted sheet
        corrupted = newXls.add_worksheet('Corrupted PDF')
        corrupted.write('A1', '')
        corrupted.write('B1', 'Nama File')
        corrupted.set_column(0,29,15)
        #close and save
        newXls.close()
    

def removeafterfs(str1, *args):
    str1 = re.sub('[.].*', '', str1)
    return str1


def removecharcom(str1, *args):
    str1 = re.sub(',', '', str1)
    str1 = re.sub('[a-zA-Z]|\n|[()]|:', '', str1)
    return str1


def removespace(str1, *args):
    str1 = re.sub(' ', '', str1)
    return str1

#PDF VALUES GETTER
def pdf_to_csv(filename):
    from io import StringIO
    from pdfminer.converter import LTChar, TextConverter
    from pdfminer.layout import LAParams
    from pdfminer.pdfdocument import PDFDocument
    from pdfminer.pdfpage import PDFPage
    from pdfminer.pdfparser import PDFParser
    from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter

    class CsvConverter(TextConverter):
        def __init__(self, *args, **kwargs):
            TextConverter.__init__(self, *args, **kwargs)

        def end_page(self, i):
            from collections import defaultdict
            lines = defaultdict(lambda: {})
            for child in self.cur_item._objs:
                if isinstance(child, LTChar):
                    (_, _, x, y) = child.bbox
                    line = lines[int(-y)]
                    line[x] = child.get_text()
                    # the line is now an unsorted dict

            for y in sorted(lines.keys()):
                line = lines[y]
                # combine close letters to form columns
                xpos = tuple(sorted(line.keys()))
                new_line = []
                temp_text = ''
                for i in range(len(xpos)-1):
                    temp_text += line[xpos[i]]
                    if xpos[i+1] - xpos[i] > 8:
                        # the 8 is representing font-width
                        # needs adjustment for your specific pdf
                        new_line.append(temp_text)
                        temp_text = ''
                # adding the last column which also manually needs the last letter
                new_line.append(temp_text+line[xpos[-1]])

                self.outfp.write(";".join(nl for nl in new_line))
                self.outfp.write("\n")

    # ... the following part of the code is a remix of the
    # convert() function in the pdfminer/tools/pdf2text module
    rsrc = PDFResourceManager()
    outfp = StringIO()
    device = CsvConverter(rsrc, outfp, laparams=LAParams())

    fp = open(filename, 'rb')
    parser = PDFParser(fp)
    doc = PDFDocument(parser)
    password = ""
    maxpages = 0
    caching = True
    pagenos = set()

    interpreter = PDFPageInterpreter(rsrc, device)

    for i, page in enumerate(PDFPage.get_pages(fp,
                                               pagenos, maxpages=maxpages,
                                               password=password, caching=caching,
                                               check_extractable=True)):
        outfp.write("START PAGE %d\n" % i)
        if page is not None:
            interpreter.process_page(page)
        outfp.write("END PAGE %d\n" % i)

    device.close()
    fp.close()

    return outfp.getvalue()

#PDF getter pemanggil
def pdf_conv():
    folder = r'/Users/gerryasrillinsandy/git/GerryAsrillinSandy/Monthly Billing Convertion Tool/PDF Source'
    print(resultDir)
    print(folder)
    chdir(resultDir)
    if path.exists('Billing Result '+ str(today) + '.xlsx'):   
        chdir(folder)
        filesProcessed = 0
        pln1files = 0
        pln2files = 0
        batamfiles = 0
        mandirifiles = 0
        pdfFiles = [f for f in listdir('.') if f.endswith('.pdf')]
        pln1collected = []
        pln2collected = []
        batamcollected = []
        mandiricollected= []
        for pdfFile in pdfFiles:
            rapih = []
            filesProcessed += 1
            print(pdfFile)
            try:
                fn = pdfFile
                result = pdf_to_csv(fn)
                lines = result.split('\n')
                for line in lines:
                    rapih.append(line.split(";"))
                csv_conv(rapih, folder, pln1collected, pln2collected, batamcollected, mandiricollected, pln1files, pln2files, batamfiles,mandirifiles)
            except Exception as e:
                chdir(resultDir)
                # chdir(resultDir)
                print("Conversion Failed: ", pdfFile, e)
                nBook2 = load_workbook('Billing Result '+ str(today) + '.xlsx', data_only=True)
                nSheet2 = nBook2['Corrupted PDF']
                rowcor = (("Move this File to another if needed:", str(pdfFile)))
                nSheet2.append(rowcor)
                nBook2.save('Billing Result '+ str(today) + '.xlsx')
                chdir(folder)
                continue
        chdir(resultDir)
        try:
            nBook = load_workbook('Billing Result '+ str(today) + '.xlsx', data_only =True)
            pln1Sheet = nBook['1st PLN']
            pln2Sheet = nBook['2nd PLN']
            batamSheet = nBook['Batam']
            mandiriSheet = nBook['Mandiri']
            pln1app = [pln1Sheet.append(y) for y in pln1collected]
            pln2app = [pln2Sheet.append(a) for a in pln2collected]
            batapp = [batamSheet.append(i) for i in batamcollected]
            manapp = [mandiriSheet.append(x) for x in mandiricollected]
            nBook.save('Billing Result '+ str(today) + '.xlsx')
        except Exception as e:
            print("Error while appending result to Excel, Error Code:", e)
            print("\n")
        print("Total PDF Processed : ", filesProcessed)
        
    else:
        print("Result file not found")


#PDF Values Setter and Appender
def csv_conv(rapih, folder,pln1collected, pln2collected, batamcollected, mandiricollected, pln1files, pln2files, batamfiles,mandirifiles):
      # batam
        idpel, rek, bul, gol, daya, faktor = "", "", "", "", "", ""
        tagbel, payment, tagnow, beban = "", "", "", ""
        bkwh, kvarhlebih, materai, latecharge, transcharge = "", "", "", "", ""
        capcharge, tampajak, trfadjust, ppjnum, angsuran = "", "", "", "", ""
        qreduction, lain, sisa = "", "", ""
        tglkwhbef, poskwhbef, tglkwhaft, poskwhaft = "", "", "", ""
        b1pemkwh, b1ruper, b1perper = "", "", ""
        b2pemkwh, b2ruper, b2perper = "", "", ""
        batnam, status, tagser, ppj = "", "", "", ""
        
        #1 PLN
        swlwbp, sklwbp, kwhlwbp = "", "", "" 
        swwbp, skwbp, kwhwbp = "", "", ""
        swkvarh, skkvarh, kvarh = "", "", ""
        tlwbp, twbp, tkvarh = "", "", ""
        tunggakan, bp, ujl, emin, totcicilan = "", "", "", "", ""
        rpttl, kompen, ppn, rpjasa, totagihan= "", "", "", "", ""
        nama, subsidi, p2tl, susulan = "", "", "", ""
        
        #2 Pln
        nama2, idpel2, jamnyala, stakhir, stawal, = "", "", "", "", ""
        pemlwbp, blwbp, bsublwbp = "", "", "" 
        pemwbp, bwbp, bsubwbp= "", "", ""
        pemkvarh, bkvarh, bsubkvarh = "", "", ""
        totpm, ptlbruto, ptlnet, ptltagih, lain2 = "", "", "", "", ""
        jumptl, totppn, ptlpersen, ptljums = "", "", "", ""
        jumtag, batas, bk, tgl2 = "", "", "", ""
        fkm = ""
        aklwbp, akwbp, awlwbp, awwbp, lwbpresetg, lwbpreset = "", "", "", "", "", ""
        
        # mandiri
        manid, manam, manrek, mangol, mandaya, manjam, mantot= "", "", "", "", "", "", ""
        manref, manblth, manstand, manadmin, mantgl, mantag = "", "", "", "", "", ""
        if 'Pelunasan' in rapih[1][0]:
            # batam
            # for d in rapih:
            #     print(d)
            batamfiles = batamfiles+1
            try:
                # batnam
                if len(rapih[2]) > 1:
                    try:
                        batnam = ''.join(rapih[2])
                        batnam = re.sub("-", "", batnam)
                    except:
                        print("Name not detected")
                else:
                    print("There is something wrong with the names")
            except:
                print("Failed to get Name")
            for row in rapih:
                for data in row:
                    if 'Nomor Rekening' in data:
                        try:
                            rek = row[-1]
                        except:
                            print("Not Found")
                    else:
                        pass
                    if 'Tarif Daya' in data:
                        try:
                            g = row[-1].split("-")
                            gol = g[0]
                            daya = g[-1]
                        except:
                            print("Error")
                    elif 'Rekening Bulan' in data:
                        try:
                            bul = row[-1]
                        except:
                            print("Error")
                    elif 'Pelunasan' in data:
                        try:
                            status = row[-1]
                        except:
                            print("Error")
                    elif 'Consumer Number' in data:
                        c = 0
                        while c < len(rapih):
                            if 'Consumer Number' in rapih[c][0]:
                                try:
                                    idpel = rapih[c+1][0]
                                    print(idpel)
                                except:
                                    print("error")
                            elif 'Tarif PPJ' in rapih[c][0]:
                                try:
                                    ppj = rapih[c+1][0]
                                except:
                                    print("error")
                            else:
                                pass
                            c += 1
                    elif 'Tagihan Sebelum' in data:
                        try:
                            tagbel = re.sub("[.|A-Za-z]", "", row[-1])
                        except:
                            print("error Tagihan Sebelumnya")
                    elif 'Pembayaran /' in data:
                        try:
                            payment = re.sub("[.|A-Za-z]", "",row[-2])
                        except:
                            print("payment error")
                    elif 'Current Balance' in data:
                        try:
                            tagnow = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("error tagihan sekarang")
                    elif 'Biaya Beban' in data:
                        if beban == "":
                            try:
                                beban = re.sub("[.|A-Za-z]", "", row[-1])
                            except:
                                print("beban biaya error")
                        else:
                            pass
                    elif 'Pemakaian KWh' in data:
                        if bkwh == "":
                            try:
                                bkwh = re.sub("[.|A-Za-z]", "",row[-1])
                            except:
                                print("Biaya KWh Error")
                    elif 'Kelebihan KVARh' in data:
                        try:
                            kvarhlebih = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("KVARh Lebih error")
                    elif 'Materai' in data:
                        try:
                            materai = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("Materai Error")
                    elif 'Keterlambatan' in data:
                        try:
                            latecharge = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("Biaya Keterlambatan not found")
                    elif 'Transformer' in data:
                        try:
                            transcharge = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("Error trans")
                    elif 'Capasitor' in data:
                        try:
                            capcharge = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("Capasitor Charge error")
                    elif 'Pertambahan Nilai' in data:
                        try:
                            tampajak = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("Tambahan Pajak Error")
                    elif 'Listrik Berkala' in data:
                        try:
                            trfadjust = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("Tarif adjustment error")
                    elif'Penerangan Jalan' in data:
                        try:
                            ppjnum = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("PPJ Number error")
                    elif 'TS' in data:
                        try:
                            angsuran = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("angsuran error")
                    elif 'Kompensasi' in data:
                        try:
                            qreduction = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("Kompensasi/Quality Reduction Error")
                    elif 'Lain' in data:
                        try:
                            lain = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("Lain Error")
                    elif 'Total Amount Due' in data:
                        try:
                            tagser = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("error")
                    elif 'Installment Balance' in data:
                        try:
                            sisa = re.sub("[.|A-Za-z]", "",row[-1])
                        except:
                            print("sisa error")
                    elif data == "KWh/KVARh":
                        c = 0
                        print(rapih[c])
                        while c < len(rapih):
                            if rapih[c][0] == "Ukur":
                                try:
                                    tglkwhbef = rapih[c+1][0]
                                except:
                                    print("tanggal kwh before empty")
                                try:
                                    poskwhbef = rapih[c+1][1]
                                except:
                                    print("posisi kwh before not found")
                                try:
                                    tglkwhaft = rapih[c+1][2]
                                except:
                                    print("Tanggal Kwh After Not found")
                                try:
                                    poskwhaft = rapih[c+1][3]
                                except:
                                    print("Pos kwh After not found")
                                try:
                                    faktor = rapih[c+1][4]
                                except:
                                    print("error faktor")
                            else:
                                pass
                            c += 1

                    elif data == "Blok I":
                        try:
                            b1pemkwh = re.sub("[.|A-Za-z]", "",row[1])
                            print(b1pemkwh)
                        except:
                            print("B1 Pemakaian KWH error")
                        try:
                            b1ruper = re.sub("[.|A-Za-z]", "",row[2])
                        except:
                            print("B1 Rupiah Per pemakaian error")
                        try:
                            b1perper = re.sub("[.|A-Za-z]", "",row[3])
                        except:
                            print("B1 Perhitungan pemakaian error")
                    elif data == "Blok II":
                        try:
                            b2pemkwh = re.sub("[.|A-Za-z]", "",row[1])
                        except:
                            print("B2 Pemakaian KWH error")
                        try:
                            b2ruper = re.sub("[.|A-Za-z]", "",row[2])
                        except:
                            print("B2 Rupiah Per pemakaian error")
                        try:
                            b2perper = re.sub("[.|A-Za-z]", "",row[3])
                        except:
                            print("B2 Perhitungan pemakaian error")
                    else:
                        pass
            batamdata = ((idpel, batnam, rek, bul, gol, daya, faktor, status, tagbel, payment, tagnow, trfadjust, beban, bkwh, kvarhlebih, materai, tampajak, latecharge, transcharge,
                          capcharge, ppj, ppjnum, angsuran, qreduction, lain, tagser, sisa, tglkwhbef, tglkwhaft, poskwhbef, poskwhaft, b1pemkwh, b1ruper, b1perper, b2pemkwh, b2ruper, b2perper ))
            batamfix = []
            for data in batamdata:
                if data =="":
                    data= "0"
                    batamfix.append(data)
                else:
                    batamfix.append(data)
            batamcollected.append(batamfix)
        elif 'MANDIRI' in rapih[1][0]:
            mandirifiles += 1
            rowcounter = 0
            # for i in rapih:
            #     print(i)
            for row in rapih:
                for data in row:
                    if 'TGL' in data:
                        try:
                            wkt = data.split()
                            mantgl = wkt[3]
                            manjam = wkt[4]
                        except:
                            print("error Tanggal")
                    elif 'NAMA' in data:
                        try:
                            manam = row[2]
                            manstand = row[-1]
                        except:
                            print("Nama not found")
                    elif 'IDPEL' in data:
                        if manid =="":
                            try:
                                manid = row[2]
                            except:
                                print("error idpel")
                        else:
                            pass
                    elif 'TARIP' in data:
                        try:
                            goldar = row[2].split("/")
                            mangol = goldar[0]
                            mandaya = goldar[1]
                        except:
                            print("error Golongan/Daya")
                    elif "TAG PLN" in data:
                        try:
                            mantag = row[-1].split()
                            mantag = removecharcom(mantag[-1])
                        except:
                            print("Tag Project error")
                    elif "NO. REF" in data:
                        if manref == "":
                            try:
                                manref = removespace(row[-1])
                            except:
                                print("No. Ref Error")
                        else:
                            pass
                    elif "BL/TH" in data:
                        try:
                            manblth = row[-1]
                        except:
                            print("BLTH Error")
                    elif "STAND METER" in data:
                        try:
                            manstand = row[2]
                            manadmin = removecharcom(removespace(data[-1]))
                        except:
                            print("Stand Meter error")
                    elif 'ADMIN BANK' in data:
                        try:
                            manadmin = row[2].split()
                            manadmin = removecharcom(manadmin[-1])
                        except:
                            print("ADMIN ERROR")
                    elif "TOTAL BAYAR" in data:
                        try:
                            mantot = row[-1].split()
                            mantot = removecharcom(mantot[-1])
                        except:
                            print("Total Bayar error")
                    elif 'CETAK ASLI' in data:
                        mandata= ((manid, manam, mantgl, manjam, mangol, mandaya, mantag, manref, manblth, manstand, manadmin, mantot))
                        # print(mandata)
                        mandiricollected.append(mandata)
                        rowcounter += 1
                        manid, manam, mangol, mandaya, manjam, mantot= "", "", "", "", "", ""
                        manref, manblth, manstand, manadmin, mantgl, mantag = "", "", "", "", "", ""
                        manfix = []
                        for datas in manfix:
                            if datas =="":
                                datas = "0"
                                manfix.append(datas)
                            else:
                                manfix.append(datas)
                        if manfix:
                            mandiricollected.append(manfix)
                        else:
                            pass
                    else:
                        pass
        elif rapih[1][0] =='PT. PERUSAHAAN LISTRIK NEGARA (PERSERO)':
            print("PLN 1")
            pln1files +=1
            lp2tl = []
            lsusulan = []
            c= 0
            while c < len(rapih):
                a= 0
                while a < len(rapih[c]):
                    if len(rapih[c]) == 1:
                        try:
                            if re.search("[0-9]", rapih[c][a]):
                                s = removecharcom(rapih[c][0])
                                if "UJL" in rapih[c+1][0]:
                                    ujl = s
                                elif "P2TL" in rapih[c+1][0]:
                                    lp2tl.append(int(s))
                                elif rapih[c+1][0]=='k':
                                    lsusulan.append(int(s))
                                else:
                                    pass
                            else:
                                pass
                        except Exception as e:
                            print("Error susulan and p2tl:", e)
                    a+=1
                c+=1
            for row in rapih:
                for data in row:
                    if "ID Pelanggan" in data and len(row)== 3:
                        try:    
                            idpel = re.sub(":", "", row[1])
                        except:
                            print("ID Pelanggan not found")
                    elif "Nama Pelanggan" in data:
                        try:
                            if len(data) >2:
                                nama = ''.join(row[1:])
                                nama = re.sub(":", "", nama)
                            elif len(data) == 1:
                                pass
                            elif len(data) == 2:
                                nama = data[1]
                            else:
                                pass
                        except:
                            print("Nama Pelanggan error")
                    elif "Subsidi*" in data:
                        try:
                            subsidi = removecharcom(row[-1])
                        except:
                            print("subsidi error")
                    elif "Total Tagihan**" in data:
                        try:
                            split = data.split(" ")
                            totagihan = removecharcom(split[-1])
                        except:
                            print("Total tagihan Error")
                    elif "Rekening" in data and "No" in data:
                        try:
                            split = data.split(" ")
                            rek = split[-1]
                        except:
                            print("Rekening Error")
                    elif "Status" in data:
                        try:
                            status = row[-1]
                        except:
                            print("Status Error")
                    elif "Golongan Tarif" in data:
                        try:
                            split= row[-1].split("/")
                            gol = re.sub(":", "", split[0])
                            daya = split[1]
                        except:
                            print("Golongan tarif error")
                    elif "Faktor Kali Meter" in data:
                        try:
                            faktor2 = re.sub(":", "", row[-1])
                        except:
                            print("Faktor Kali Meter Error")
                    elif "Tunggakan Bulan Sebelu" in data:
                        try:
                            tunggakan= removecharcom(row[-1])
                        except:
                            print("Tunggakan Bulan Sebelumnya Not found")
                    elif "Biaya Penyambungan" in data:
                        try:
                            bp = removecharcom(row[-1])
                        except:
                            print("BP error")
                    elif "UJL" in data and len(row) == 3:
                        try:
                            ujl = removecharcom(row[-1])
                        except:
                            print("UJL Error")
                    elif "P2TL" in data and len(row) == 3:
                        try:
                            lp2tl.append(int(removecharcom(row[-1])))
                        except:
                            print("P2TL error")
                    elif "Susulan" in data and len(row) == 4:
                        try:
                            lsusulan.append(int(removecharcom(row[-1])))
                        except Exception as e:
                            print("Susulan Error:", e)
                    elif "Angsuran Lainnya" in data:
                        try:
                            lain = removecharcom(row[-1])
                        except:
                            print("Angsuran Lainnya Error")
                    elif "Biaya Beban" in data:
                        try:
                            emin = removecharcom(row[-1])
                        except:
                            print("Biaya Beban Error")
                    elif "Stand Akhir" in data:
                        sta = []
                        try:
                            for data2 in row:
                                if "Stand Akhir" in data2:
                                    pass
                                elif "PLN 123" in data2:
                                    pass
                                elif "pln123" in data2:
                                    pass
                                else:
                                    data3 = removespace(removecharcom(data2))
                                    if data3 == "":
                                        pass
                                    else:
                                        sta.append(data3)
                            if len(sta) == 1:
                                sklwbp = sta[0]
                            elif len(sta) == 2 :
                                sklwbp = sta[0]
                                skwbp = sta[1]
                            elif len(sta) == 3:
                                sklwbp = sta[0]
                                skwbp = sta[1]
                                skkvarh = sta[2]
                            else:
                                print("More than 3 Stand akhir")
                        except Exception as e:
                            print("stand Akhir error:", e)
                    elif "Stand Awal" in data:
                        sta = []
                        try:
                            for data2 in row:
                                if "Stand Awal" in data2:
                                    pass
                                elif "PLN 123" in data2:
                                    pass
                                elif "pln123" in data2:
                                    pass
                                else:
                                    data3 = removespace(removecharcom(data2))
                                    if data3 == "":
                                        pass
                                    else:
                                        sta.append(data3)
                            if len(sta) == 1:
                                swlwbp = sta[0]
                            elif len(sta) == 2 :
                                swlwbp = sta[0]
                                swwbp = sta[1]
                            elif len(sta) == 3:
                                swlwbp = sta[0]
                                swwbp = sta[1]
                                swkvarh = sta[2]
                            else:
                                print("More than 3 Stand Aawal")
                        except Exception as e:
                            print("stand Awal error", e)
                    elif data == "Wh L":
                        try:
                            k = row[2].split(":")
                            kwhlwbp = removespace(removecharcom(k[-1]))
                            k2 = row[0].split(":")
                            tlwbp = removespace(removecharcom(k2[-1]))
                        except Exception as e:
                            print("Tarif LWBP Error", e)
                    elif data == "Wh ":
                        try:
                            k = row[2].split(":")
                            kwhwbp = removespace(removecharcom(k[-1]))
                            k2 = row[0].split(":")
                            twbp= removespace(removecharcom(k2[-1]))
                        except Exception as e:
                            print("Tarif WBP error:", e)
                    elif data == "Tarif kVArh":
                        try:
                            k = row[-1].split(" ")
                            tkvarh = removespace(removecharcom(k[-1]))
                            k2 = row[0].split(":")
                            kvarh = removespace(removecharcom(k2[-1]))
                        except:
                            print("Tarif Kvarh Error")
                    elif "TTL Terpakai" in data:
                        try:
                            rpttl = removecharcom(row[-1])
                        except:
                            print("RPTTL Error")
                    elif data == "minus Ko":
                        try:
                            kompen = removecharcom(row[-1])
                        except:
                            print("Kompensasi Error")
                    elif "PPN*" in data:
                        try:
                            ppn = removecharcom(row[-1])
                        except:
                            print("PPN Error")
                    elif "PPJ*" in data:
                        try:
                            ppj = removecharcom(row[-1])
                        except:
                            print("PPJ error")
                    elif "dan Keandalan" in data:
                        try:
                            rpjasa = removecharcom(row[-1])
                        except:
                            print("Rupiah Jasa Not Found")
                    else:
                        pass
            try:
                susulan = sum(lsusulan)
            except Exception as e:
                print("Susulan Error", e)
            try:
                p2tl = sum(lp2tl)
            except Exception as e:
                print("P2TL Error:", e)
            print("Susulan final:",susulan)
            print("P2TL final:", p2tl)
            totcicilan = p2tl + susulan
            pln1data = ((idpel, rek, nama, subsidi, gol, daya, faktor2,swlwbp, sklwbp, kwhlwbp, swwbp, skwbp, kwhwbp, swkvarh, skkvarh, kvarh, tlwbp, twbp, tkvarh, tunggakan, bp, susulan, p2tl, ujl, lain, emin, totcicilan, rpttl, kompen, ppj, ppn, rpjasa, totagihan, status))     
            pln1fix = []
            for data in pln1data:
                if data =="":
                    data = "0"
                    pln1fix.append(data)
                else:
                    pln1fix.append(data)
            
            pln1collected.append(pln1fix)
            idpel, rek, nama, subsidi, gol, daya, faktor2,swlwbp, sklwbp, kwhlwbp, swwbp, skwbp, kwhwbp, swkvarh, skkvarh, kvarh, tlwbp, twbp, tkvarh, tunggakan, bp, susulan, p2tl, ujl, lain, emin, totcicilan, rpttl, kompen, ppj, ppn, rpjasa, totagihan, status = "", "", "", "", "","", "", "", "", "","", "", "", "", "" ,"", "", "", "", "","", "", "", "", "", "", "", "", "", "", "", "", "", ""
            # c = 0
            # while c < len(rapih):
        elif 'UID' in rapih[1][0] or 'UIW' in rapih[1][0]:
            print("PLN 2")
            # for row in rapih:
            #     print(row)
            pln2files += 1
            for row in rapih:
                for data in row:
                    if 'Kepada Yth' in data:
                        try:
                            nama2 = row[1]
                            idpel2 = re.sub(":", "", row[-1])
                        except:
                            print("Name error")
                    elif 'Rekening Bulan' in data:
                        try:
                            bul = re.sub(":", "",row[-1])
                        except:
                            print("bul error")
                    elif data == "Tarif / Daya":
                        try:
                            split = row[-1].split("/")
                            gol = removespace(split[0])
                            daya = removespace(split[1])
                        except:
                            print("Golongan and daya error")
                    elif str(data) == "FKM kWh/kVarh":
                        try:
                            fkm = re.sub(":", "", row[-1])
                        except:
                            print("FKM error")
                    elif "Jam Nyala" in data:
                        try:
                            jamnyala = re.sub(":", "", row[-1])
                        except:
                            print("Jam error")
                    elif "St Akhir" in data:
                        try:
                            stakhir = removecharcom(row[1])
                        except:
                            print("St Akhir error")
                        try:
                            aklwbp = removecharcom(row[2])
                        except:
                            print("LWBP Akhir error")
                        try:
                            akwbp = removecharcom(row[3])
                        except:
                            print("WBP Akhir Error")
                    elif "St Awal" in data:
                        try:
                            stawal = removecharcom(row[1])
                        except:
                            print("St Awal error")
                        try:
                            awlwbp = removecharcom(row[2])
                        except:
                            print("LWBP Awal error")
                        try:
                            awwbp = removecharcom(row[3])
                        except:
                            print("WBP Awal Error")
                    elif "St Meter Pasang" in data:
                        try:
                            lwbpresetg = removecharcom(row[1])
                        except:
                            print("Tanggal Meter Reset not found")
                        try:
                            lwbpreset = removecharcom(row[2])
                        except:
                            print("LWBP")
                    elif "Rupiah PTL Bruto" in data:
                        try:
                            ptlbruto = removecharcom(row[-1])
                        except:
                            print("ptlbruto error")
                    elif "Kompensasi" in data:
                        try:
                            kompen = removecharcom(row[-1])
                        except:
                            print("Kompensasi error")
                    elif 'PTL Netto' in data:
                        try:
                            ptlnet = removecharcom(row[-1])
                        except:
                            print("PTL Netto error")
                    elif "Jumlah Rupiah Pemakaian Tenaga Listrik (PTL) yang ditagihkan" in data:
                        try:
                            ptltagih = removecharcom(row[-1])
                        except:
                            print("PTL Tagihan error")
                    elif "Lainnya" in data:
                        try:
                            lain2 = removecharcom(row[-1])
                        except:
                            print("Lainnya")
                    elif data == "Jumlah Rupiah Pemakaian Tenaga Listrik (PTL) (6+7)":
                        try:
                            jumptl = removecharcom(row[-1])
                        except:
                            print("Jumlah PTL Error")
                    elif data == "9.":
                        try:
                            ppn = removecharcom(row[-1])
                        except:
                            print("PPN Error")
                    elif data == "Total Penyerahan Listrik":
                        try:
                            totppn = removecharcom(row[-1])
                        except:
                            print("Total Penyerahan PPN Error")
                    elif "Jumlah Tagihan" in data:
                        try:
                            jumtag = removecharcom(row[-1])
                        except:
                            print("Jumlah Tagihan Error")
                    elif "Batas Akhir Masa Bayar" in data:
                        try:
                            batas = data.split(" ")
                            batas = str(batas[-3]) + " " + str(batas[-2]) + " " + str(batas[-1]) 
                        except:
                            print("Batas Akhir Masa Bayar error")
                    elif data == "Status":
                        try:
                            status = re.sub(":", "", row[-1])
                        except:
                            print("Status error")
                    elif data == "Tanggal Bayar":
                        try:
                            tgl2 = re.sub(":", "", row[-1])
                        except:
                            print("Tanggal Bayar error")
                    elif "BK" in data:
                        try:
                            bk = removecharcom(re.sub(".", "", row[-1]))
                        except:
                            print("Biaya Keterlambatan error") 
                    else:
                        pass
            c= 0
            while c < len(rapih):
                a = 0
                while a <len(rapih[c]):
                    if rapih[c][a] == "Pemk kWh":
                        try:
                            totpm = removecharcom(rapih[c+1][1])
                            pemlwbp = removecharcom(rapih[c+1][0])
                            blwbp = removecharcom(rapih[c+2][1])
                            bsublwbp = removecharcom(rapih[c+2][2])
                        except:
                            print("Total and pemakaian LWBP error")
                        try: 
                            pemwbp = removecharcom(rapih[c+2][3])
                            bwbp = removecharcom(rapih[c+2][4])
                            bsubwbp = removecharcom(rapih[c+2][5])
                        except:             
                            print("WBP Part Problem")
                        try:
                            pemkvarh = removecharcom(rapih[c+2][6])
                            bkvarh = removecharcom(rapih[c+2][7])
                            bsubkvarh = removecharcom(rapih[c+2][8])
                        except:
                            print("KVarh error")
                    elif str(rapih[c][a]) == "Pendapatan Biaya Beban":
                        try:
                            beban = removecharcom(rapih[c-1][-1])
                        except:
                            print("Biaya Beban Error")
                    elif rapih[c][a] == "II":
                        if rapih[c][-1]== "0":
                            pass
                        else:
                            try:
                                ptlpersen = rapih[c+1][0]
                                ptljums = removecharcom(rapih[c+1][-1])
                            except:
                                print("PTL Persen and Jumlah Crazy")
                    else:
                        pass
                    a +=1
                c+=1
            pln2data = ((nama2, idpel2, bul, gol, daya, fkm, jamnyala, stakhir, aklwbp, akwbp, stawal, awlwbp, awwbp, lwbpresetg, lwbpreset, beban, pemlwbp, blwbp, bsublwbp, pemwbp, bwbp, bsubwbp, pemkvarh, bkvarh, bsubkvarh,  totpm, ptlbruto, kompen, ptlnet, ptltagih, lain2, jumptl, ppn, totppn, ptlpersen, ptljums, jumtag, batas, status, tgl2, bk))                    
            pln2fix = []
            for data in pln2data:
                if data =="":
                    data = "0"
                    pln2fix.append(data)
                else:
                    pln2fix.append(data)
                
            pln2collected.append(pln2fix)

            nama2, idpel2, bul, gol, daya, fkm, jamnyala, stakhir, stawal, beban, pemlwbp, blwbp, bsublwbp, pemwbp, bwbp, bsubwbp, pemkvarh, bkvarh, bsubkvarh, ppn, totpm, ptlbruto, kompen, ptlnet, ptltagih, lain2, jumptl, totppn, ptlpersen, ptljums, jumtag, batas, status, tgl2, bk = "", "", "", "", "","", "", "", "", "","", "", "", "", "" ,"", "", "", "", "","", "", "", "", "", "", "", "", "", "", "", "", "", "", ""
        else:
            print("New Format detected")  



createResult()
pdf_conv()